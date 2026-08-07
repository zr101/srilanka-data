"""Shared helpers for CBSL's statistical-table workbooks.

Three site behaviours shape everything here (see srilankamonitor
docs/CBSL-STATISTICS.md for the full census):

1. Workbook filenames embed a mutable date stamp (`table2.02_20260731_e.xlsx`)
   and there is no "latest" alias, so the current URL must be discovered from
   the listing page on every run. The stamp changing *is* the update signal.
2. **Missing files return HTTP 200 with an HTML error body.** Status codes are
   useless; only the `PK` zip magic distinguishes a real workbook.
3. Sheet names and header text drift between vintages (year ranges roll over,
   `RDT (June 2026)` renames monthly), so callers must anchor on header *text*,
   never on fixed row/column offsets or sheet indexes.
"""

import io
import re
import unicodedata

import openpyxl

from .http import get

SHEETS_BASE = "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/statistics/sheets"

MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]
_MONTH_INDEX = {m[:3]: i + 1 for i, m in enumerate(MONTHS)}


def norm(value) -> str:
    """Collapse a cell/header to comparable text: NBSP, newlines and repeated
    spaces all vary between vintages of the same table."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def key(value) -> str:
    """norm() plus case- and punctuation-folding, for matching header labels."""
    return re.sub(r"[^a-z0-9]+", "", norm(value).lower())


def num(value) -> float | None:
    """Parse a numeric cell.

    CBSL uses 'n.a', '-' and '...' for missing, '(1.2)' for negatives, and
    appends roman-numeral footnote markers to some figures ('6,458 (II)').
    A trailing parenthetical is a footnote when its contents aren't numeric,
    and a negation when they are.
    """
    if isinstance(value, (int, float)):
        return float(value)
    text = norm(value)
    if not text:
        return None
    text = re.sub(r"\s*\([^)0-9]*\)$", "", text)  # drop footnote markers
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "")
    if not re.fullmatch(r"-?\d*\.?\d+", text):
        return None  # n.a, -, ..., footnote markers
    result = float(text)
    return -result if negative else result


def month_of(text: str) -> int | None:
    """Month number from a full or abbreviated English month name."""
    return _MONTH_INDEX.get(norm(text).lower()[:3])


def period_of(text: str) -> str | None:
    """'2025 June' / 'June 2025' → '2025-06'."""
    text = norm(text)
    year = re.search(r"(19|20)\d{2}", text)
    if not year:
        return None
    month = None
    for token in re.findall(r"[A-Za-z]+", text):
        month = month_of(token)
        if month:
            break
    return f"{year.group(0)}-{month:02d}" if month else None


def month_stamp(value) -> str | None:
    """'YYYY-MM' from a period header cell, whatever type it arrived as.

    The trade workbooks start their header as real date cells and switch to
    text partway along ("Jan-26 (b)"), so matching only datetimes silently
    truncates the series years before the present.
    """
    if hasattr(value, "year") and hasattr(value, "month"):
        return f"{value.year}-{value.month:02d}"
    text = norm(value)
    if not text:
        return None
    month = next((month_of(t) for t in re.findall(r"[A-Za-z]+", text) if month_of(t)), None)
    if month is None:
        return None
    year_match = re.search(r"(?<!\d)(\d{4}|\d{2})(?!\d)", re.sub(r"[A-Za-z]+", " ", text))
    if not year_match:
        return None
    year = int(year_match.group(1))
    if year < 100:
        year += 2000 if year < 70 else 1900
    return f"{year}-{month:02d}"


def month_columns(ws, row: int, start_col: int = 2) -> dict[str, int]:
    """{YYYY-MM: column} for a header row of period labels of any type."""
    out: dict[str, int] = {}
    for col in range(start_col, ws.max_column + 1):
        stamp = month_stamp(ws.cell(row, col).value)
        if stamp and stamp not in out:
            out[stamp] = col
    return out


def get_xlsx(url: str) -> bytes | None:
    """Download a workbook, returning None when CBSL serves its HTML
    soft-404 (HTTP 200 with an `<!` body) or a real 404."""
    res = get(url)
    if res.status_code == 404:
        return None
    if not res.content.startswith(b"PK"):
        return None
    return res.content


def load(payload: bytes):
    """Open a workbook's cached values. Not read_only: merged-cell ranges and
    random access are both needed, and these books are small enough."""
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=True)


def discover(listing_url: str) -> dict[str, str]:
    """Scrape a statistical-tables listing page into {link text: absolute URL}.

    Anchors wrap the label in nested markup and the href may be relative, so
    both sides need cleaning. Later duplicates lose to the first occurrence,
    which is the linked (current) vintage.
    """
    html = get(listing_url).text
    found: dict[str, str] = {}
    for match in re.finditer(
        r'<a[^>]+href="([^"]+\.(?:xlsx|xls))"[^>]*>(.*?)</a>', html, re.S | re.I
    ):
        url, label = match.group(1), norm(re.sub(r"<[^>]+>", " ", match.group(2)))
        if url.startswith("/"):
            url = "https://www.cbsl.gov.lk" + url
        if label and label not in found:
            found[label] = url
    return found


def pick(listing: dict[str, str], *needles: str) -> tuple[str, str] | None:
    """Find the (label, url) whose label contains every needle, case-insensitively.

    Labels carry year ranges that change ("(2006 to Latest)"), so callers match
    on the stable words only.
    """
    wanted = [n.lower() for n in needles]
    for label, url in listing.items():
        low = label.lower()
        if all(n in low for n in wanted):
            return label, url
    return None


def stamp_of(url: str) -> str | None:
    """The YYYYMMDD vintage stamp embedded in a workbook filename, if present."""
    match = re.search(r"(20\d{6})", url)
    return match.group(1) if match else None


def sheet(wb, *needles: str):
    """Worksheet whose name contains every needle. Sheet names drift between
    vintages ('2.02 In USD 2007-2025' → '…2007-2026'), so never index by
    position or match a full name."""
    wanted = [n.lower() for n in needles]
    for name in wb.sheetnames:
        low = norm(name).lower()
        if all(n in low for n in wanted):
            return wb[name]
    return None


def find_row(ws, *needles: str, column: int = 1, limit: int = 60, start: int = 1) -> int | None:
    """Row number whose cell in `column` matches every needle (folded).

    `start` matters when a sheet's title repeats a data row's label — the IIP
    workbook heads itself "Index of Industrial Production (2015=100)" and also
    carries an "Index of Industrial Production" total row further down.
    """
    wanted = [key(n) for n in needles]
    for row in range(start, min(ws.max_row, limit) + 1):
        cell = key(ws.cell(row, column).value)
        if cell and all(n in cell for n in wanted):
            return row
    return None


def header_columns(ws, row: int) -> dict[str, int]:
    """{folded header text: column} for one header row, first occurrence wins."""
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        label = key(ws.cell(row, col).value)
        if label and label not in out:
            out[label] = col
    return out


def forward_fill(ws, row: int, last_col: int | None = None) -> list:
    """Read a header row, carrying each value across the blank cells that a
    merged span leaves behind (openpyxl reports the value only in the anchor)."""
    values, current = [], None
    for col in range(1, (last_col or ws.max_column) + 1):
        raw = ws.cell(row, col).value
        if raw is not None and norm(raw):
            current = raw
        values.append(current)
    return values


def leaf_columns(ws, group_row: int, sub_row: int, start_col: int) -> dict[str, int]:
    """{slug: column} for a two-tier header.

    Group labels are merged across their sub-columns, so they are forward-filled;
    a column with no sub-label is named after its group alone (single-level
    columns like "Year" or "Grants" sit in the same header block as two-level
    ones). Duplicate slugs keep their first column.
    """
    groups = forward_fill(ws, group_row)
    out: dict[str, int] = {}
    for col in range(start_col, ws.max_column + 1):
        group = key(groups[col - 1]) if groups[col - 1] else ""
        sub = key(ws.cell(sub_row, col).value) if sub_row != group_row else ""
        slug = f"{group}_{sub}" if group and sub else (group or sub)
        if slug and slug not in out:
            out[slug] = col
    return out


def periods_down(ws, start_row: int, mode: str, year_col: int, month_col: int | None = None):
    """Yield (row, period) reading the period column(s) downward.

    CBSL encodes periods three ways in these tables and the year is written
    only on its first row in every one of them, so it is carried forward:
      annual   — "1990" in year_col
      split    — year in year_col, month name in month_col
      inline   — "2003 Jan" then bare "Feb" in year_col
      datetime — a real date cell in year_col (the monetary survey)
    """
    year = None
    for row in range(start_row, ws.max_row + 1):
        if mode == "datetime":
            stamp = ws.cell(row, year_col).value
            if hasattr(stamp, "year"):
                yield row, f"{stamp.year}-{stamp.month:02d}"
            continue
        raw = norm(ws.cell(row, year_col).value)
        if mode == "annual":
            found = re.fullmatch(r"((?:19|20)\d{2})[a-z ()*]*", raw.lower())
            if found:
                yield row, found.group(1)
            continue
        if mode == "split":
            if re.match(r"(19|20)\d{2}", raw):
                year = raw[:4]
            month = month_of(ws.cell(row, month_col).value)
            if year and month:
                yield row, f"{year}-{month:02d}"
            continue
        found = re.match(r"((?:19|20)\d{2})", raw)
        if found:
            year = found.group(1)
        month = month_of(raw.split()[-1]) if raw else None
        if year and month:
            yield row, f"{year}-{month:02d}"


def series(points: dict[str, float | None], keep: int | None = None) -> list[dict]:
    """{period: value} → sorted [{t, v}] with missing periods dropped, matching
    the dashboard's MacroSeries convention (oldest first).

    Values are rounded to 4dp: these workbooks carry full float expansions
    (243.58015872545985 for a USD-million figure), which are meaningless at
    this precision and triple the size of the latest.json the dashboard pulls.
    `keep` trims to the most recent N points.
    """
    out = [{"t": t, "v": round(v, 4)} for t, v in sorted(points.items()) if v is not None]
    return out[-keep:] if keep else out
